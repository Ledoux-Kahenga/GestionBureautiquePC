"""
Utilitaires de sauvegarde et restauration de la base de données
"""
import os
import shutil
import json
import sqlite3
from datetime import datetime
from config import DATABASE_PATH, BASE_DIR


class BackupManager:
    """Gestionnaire de sauvegardes de la base de données"""
    
    def __init__(self):
        self.backup_dir = os.path.join(BASE_DIR, "backups")
        self.settings_file = os.path.join(BASE_DIR, "settings.json")
        self._ensure_backup_dir()
    
    def _ensure_backup_dir(self):
        """S'assurer que le dossier de sauvegardes existe"""
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
    
    def get_backup_path(self):
        """Retourne le chemin du dossier de sauvegardes"""
        return self.backup_dir
    
    def set_backup_path(self, path):
        """Définir un nouveau chemin pour les sauvegardes"""
        if os.path.exists(path):
            self.backup_dir = path
            self._save_settings()
            return True
        return False
    
    def _save_settings(self):
        """Sauvegarder les paramètres"""
        settings = self.load_settings()
        settings['backup_path'] = self.backup_dir
        with open(self.settings_file, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=4, ensure_ascii=False)
    
    def load_settings(self):
        """Charger les paramètres depuis le fichier JSON"""
        default_settings = {
            'backup_path': self.backup_dir,
            'auto_backup': True,
            'backup_on_close': True,
            'max_backups': 10,
            'theme': 'light',
            'pin_enabled': False,
            'pin_code': None,
            'cloture_auto_time': '23:59',
            'notify_before_cloture': True,
            'font_scale': 0
        }
        
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    saved_settings = json.load(f)
                    # Fusionner avec les valeurs par défaut
                    default_settings.update(saved_settings)
            except (json.JSONDecodeError, IOError):
                pass
        
        return default_settings
    
    def save_settings(self, settings):
        """Sauvegarder les paramètres"""
        with open(self.settings_file, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=4, ensure_ascii=False)
    
    def create_backup(self, description=""):
        """Créer une sauvegarde de la base de données"""
        self._ensure_backup_dir()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}.db"
        backup_path = os.path.join(self.backup_dir, backup_name)
        
        try:
            # Copier la base de données
            shutil.copy2(DATABASE_PATH, backup_path)
            
            # Créer un fichier de métadonnées
            metadata = {
                'timestamp': datetime.now().isoformat(),
                'description': description,
                'original_path': DATABASE_PATH,
                'size': os.path.getsize(backup_path)
            }
            
            metadata_path = backup_path.replace('.db', '_metadata.json')
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=4, ensure_ascii=False)
            
            # Nettoyer les anciennes sauvegardes
            self._cleanup_old_backups()
            
            return True, backup_path
        except Exception as e:
            return False, str(e)
    
    def _cleanup_old_backups(self):
        """Supprimer les sauvegardes les plus anciennes si le maximum est atteint"""
        settings = self.load_settings()
        max_backups = settings.get('max_backups', 10)
        
        backups = self.list_backups()
        if len(backups) > max_backups:
            # Trier par date (plus ancien en premier)
            backups_sorted = sorted(backups, key=lambda x: x['timestamp'])
            
            # Supprimer les plus anciennes
            for backup in backups_sorted[:-max_backups]:
                self.delete_backup(backup['path'])
    
    def list_backups(self):
        """Lister toutes les sauvegardes disponibles"""
        self._ensure_backup_dir()
        backups = []
        
        for filename in os.listdir(self.backup_dir):
            if filename.endswith('.db'):
                filepath = os.path.join(self.backup_dir, filename)
                metadata_path = filepath.replace('.db', '_metadata.json')
                
                backup_info = {
                    'filename': filename,
                    'path': filepath,
                    'size': os.path.getsize(filepath),
                    'timestamp': datetime.fromtimestamp(os.path.getmtime(filepath)),
                    'description': ''
                }
                
                # Charger les métadonnées si disponibles
                if os.path.exists(metadata_path):
                    try:
                        with open(metadata_path, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                            backup_info['description'] = metadata.get('description', '')
                            backup_info['timestamp'] = datetime.fromisoformat(metadata.get('timestamp', backup_info['timestamp'].isoformat()))
                    except (json.JSONDecodeError, IOError):
                        pass
                
                backups.append(backup_info)
        
        # Trier par date décroissante
        backups.sort(key=lambda x: x['timestamp'], reverse=True)
        return backups
    
    def restore_backup(self, backup_path):
        """Restaurer une sauvegarde"""
        if not os.path.exists(backup_path):
            return False, "Le fichier de sauvegarde n'existe pas"
        
        try:
            # Créer une sauvegarde de sécurité avant restauration
            self.create_backup("Sauvegarde avant restauration")
            
            # Restaurer
            shutil.copy2(backup_path, DATABASE_PATH)
            
            return True, "Restauration réussie"
        except Exception as e:
            return False, str(e)
    
    def delete_backup(self, backup_path):
        """Supprimer une sauvegarde"""
        try:
            if os.path.exists(backup_path):
                os.remove(backup_path)
                
                # Supprimer aussi les métadonnées
                metadata_path = backup_path.replace('.db', '_metadata.json')
                if os.path.exists(metadata_path):
                    os.remove(metadata_path)
                
                return True
        except Exception as e:
            print(f"Erreur lors de la suppression: {e}")
        return False
    
    def export_data(self, export_path, format_type='json'):
        """Exporter toutes les données"""
        try:
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            
            if format_type == 'json':
                return self._export_to_json(cursor, export_path)
            elif format_type == 'csv':
                return self._export_to_csv(cursor, export_path)
            elif format_type == 'excel':
                return self._export_to_excel(cursor, export_path)
            else:
                return False, "Format non supporté"
                
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def _export_to_json(self, cursor, export_path):
        """Exporter en JSON"""
        data = {
            'export_date': datetime.now().isoformat(),
            'transactions': [],
            'rapports': []
        }
        
        # Exporter les transactions
        cursor.execute('SELECT * FROM transactions')
        columns = [description[0] for description in cursor.description]
        for row in cursor.fetchall():
            data['transactions'].append(dict(zip(columns, row)))
        
        # Exporter les rapports
        cursor.execute('SELECT * FROM rapports_journaliers')
        columns = [description[0] for description in cursor.description]
        for row in cursor.fetchall():
            data['rapports'].append(dict(zip(columns, row)))
        
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
        return True, export_path
    
    def _export_to_csv(self, cursor, export_path):
        """Exporter en CSV"""
        import csv
        
        # Créer un dossier pour les fichiers CSV
        csv_dir = export_path.replace('.csv', '_export')
        if not os.path.exists(csv_dir):
            os.makedirs(csv_dir)
        
        # Exporter les transactions
        cursor.execute('SELECT * FROM transactions')
        columns = [description[0] for description in cursor.description]
        
        transactions_path = os.path.join(csv_dir, 'transactions.csv')
        with open(transactions_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(cursor.fetchall())
        
        # Exporter les rapports
        cursor.execute('SELECT * FROM rapports_journaliers')
        columns = [description[0] for description in cursor.description]
        
        rapports_path = os.path.join(csv_dir, 'rapports.csv')
        with open(rapports_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(cursor.fetchall())
        
        return True, csv_dir
    
    def _export_to_excel(self, cursor, export_path):
        """Exporter en Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return False, "Le module openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
        
        workbook = openpyxl.Workbook()
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Feuille Transactions
        ws_trans = workbook.active
        ws_trans.title = "Transactions"
        
        cursor.execute('SELECT * FROM transactions ORDER BY date DESC, created_at DESC')
        columns = [description[0] for description in cursor.description]
        
        # En-têtes
        for col, header in enumerate(columns, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_num, row_data in enumerate(cursor.fetchall(), 2):
            for col, value in enumerate(row_data, 1):
                cell = ws_trans.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
        
        # Ajuster les largeurs de colonnes
        for col in ws_trans.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_trans.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Feuille Rapports
        ws_rapp = workbook.create_sheet("Rapports")
        
        cursor.execute('SELECT * FROM rapports_journaliers ORDER BY date DESC')
        columns = [description[0] for description in cursor.description]
        
        # En-têtes
        for col, header in enumerate(columns, 1):
            cell = ws_rapp.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_num, row_data in enumerate(cursor.fetchall(), 2):
            for col, value in enumerate(row_data, 1):
                cell = ws_rapp.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
        
        # Ajuster les largeurs
        for col in ws_rapp.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_rapp.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        workbook.save(export_path)
        return True, export_path
    
    def import_data(self, import_path, merge=False):
        """Importer des données depuis un fichier JSON"""
        if not os.path.exists(import_path):
            return False, "Le fichier n'existe pas"
        
        try:
            # Créer une sauvegarde avant import
            self.create_backup("Sauvegarde avant import")
            
            with open(import_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            
            if not merge:
                # Vider les tables existantes
                cursor.execute('DELETE FROM transactions')
                cursor.execute('DELETE FROM rapports_journaliers')
            
            # Importer les transactions
            for trans in data.get('transactions', []):
                cursor.execute('''
                    INSERT OR REPLACE INTO transactions 
                    (id, type, montant, description, date, created_at, type_depense)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    trans.get('id'),
                    trans.get('type'),
                    trans.get('montant'),
                    trans.get('description'),
                    trans.get('date'),
                    trans.get('created_at'),
                    trans.get('type_depense', 'normale')
                ))
            
            # Importer les rapports
            for rapport in data.get('rapports', []):
                cursor.execute('''
                    INSERT OR REPLACE INTO rapports_journaliers 
                    (id, date, cloture, cloture_at)
                    VALUES (?, ?, ?, ?)
                ''', (
                    rapport.get('id'),
                    rapport.get('date'),
                    rapport.get('cloture'),
                    rapport.get('cloture_at')
                ))
            
            conn.commit()
            conn.close()
            
            return True, f"Import réussi: {len(data.get('transactions', []))} transactions, {len(data.get('rapports', []))} rapports"
        
        except json.JSONDecodeError:
            return False, "Le fichier JSON est invalide"
        except Exception as e:
            return False, str(e)
